"""Build the aluminium-part weighing sheet: one row per part, with its own picture.

Some parts were 3D printed instead of machined, so the CAD mass at 6061 density is only a
reference and the real ones have to be put on a scale. This writes the workbook that survey
is done on - link, picture, name, CAD mass, and an empty column to write the measured mass
into - plus a per-link roll-up that fills itself in as the measurements arrive.

Measured values come from `tools/robot_model/alu_parts_measured.json` when it exists - the handwritten masses read
off photos of the printed parts, matched to CAD parts - so the sheet can be regenerated
without losing them. Each entry carries a confidence, shown as the cell colour: green is
unambiguous, yellow fits but the A/B pair-mate or side could not be proven from the photo,
orange means the reading itself or the part is uncertain and wants a check against the part
in hand. Anything the photos could not resolve is left blank with a note, never guessed.

Usage: alu_parts_xlsx.py [out.xlsx]   (mjlab .venv python)
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = os.environ.get('ALU_PARTS_DIR', '/home/syaro/pyg_fea/fusion/alu_parts')
DEFAULT_OUT = ('/home/syaro/MikuchanRemote/Human-Pygmalion/docs/'
               'aluminium_parts_3dprint_masses.xlsx')
IMG_W, IMG_H = 156, 117                      # px in the sheet
ROW_H = 92                                   # points
FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='2F3E50')
INPUT_FILL = PatternFill('solid', fgColor='FFF3B0')      # cells the user fills in
NOTE_FILL = PatternFill('solid', fgColor='F2F4F7')
# measured-value confidence -> fill. PLA at 100 % infill is 1.24/2.70 = 0.46 of the aluminium
# CAD mass, so a ratio above that can only be a misread or a misassignment (red font on F).
CONF_FILL = {'high': PatternFill('solid', fgColor='C8F0C8'),
             'med': PatternFill('solid', fgColor='FFF3B0'),
             'low': PatternFill('solid', fgColor='FFD4A8')}
CONF_KO = {'high': '확실', 'med': '보통', 'low': '불확실'}
ALU_FILL = PatternFill('solid', fgColor='D9DEE5')       # machined aluminium, not printed
PLA_CEILING = 1.24 / 2.70
THIN = Side(style='thin', color='BFC6CF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
# parts that are NOT machined one-offs, so weighing them proves nothing about printing
STOCK = ('DR2020', 'DF2020')


def find_part(meta, occ, body):
    """The index row for a measured part: by body name when that is unique, else (occ, body).

    Occurrence names drift between CAD revisions (HipRoll2Yaw is spelled PipRoll2Yaw in the
    8/16 file) while body names do not, so the body name is the stable key.
    """
    hits = [r for r in meta if r['body'] == body]
    if len(hits) == 1:
        return hits[0]
    hits = [r for r in hits if r['occ'].split(':')[0] == occ]
    return hits[0] if len(hits) == 1 else None


def note_for(r):
    if r['group'].startswith('NoSim'):
        return '설계 참조용 (NoSim) — 실물 아님'
    if any(s in r['occ'] for s in STOCK):
        return '2020 알루미늄 압출 프로파일 (구매품)'
    if r['group'] == 'CenterPin_RS03':
        return 'ø4 핀 — 출력 대상 아닐 수 있음'
    return ''


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_OUT
    meta = json.load(open(f'{SRC}/index.json'))
    meta.sort(key=lambda r: (r['link'], -r['mass_g']))
    measured, alu, mj = {}, {}, {}
    mpath = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         'alu_parts_measured.json')
    if os.path.exists(mpath):
        mj = json.load(open(mpath))
        absent = []
        for e in mj['entries']:
            row = find_part(meta, e['occ'], e['body'])
            if row is None:
                absent.append(e['body'])
                continue
            measured[row['path']] = e
        for e in mj.get('aluminium', []):
            row = find_part(meta, e['occ'], e['body'])
            if row is not None:
                alu[row['path']] = e
        if absent:
            print(f'measured parts with no row in this index (not in this CAD revision): {absent}')

    wb = Workbook()
    ws = wb.active
    ws.title = '알루미늄 부품'

    ws['A1'] = 'Pygmalion 알루미늄 부품 — 3D 프린트 실측 질량 조사'
    ws['A1'].font = Font(name=FONT, size=14, bold=True)
    ws['A2'] = ('노란 칸(E열)에만 입력하세요. 저울로 잰 값을 그램 숫자로 — 예: 78.4  ·  '
                'D열은 CAD 형상에 알루미늄 6061(2.70 g/cm³)을 적용한 값입니다.')
    ws['A2'].font = Font(name=FONT, size=10, color='555F6D')
    ref = mj.get('reference', {}) if measured else {}
    src_doc = (f"{ref['doc']} v{ref['version']} ({ref.get('saved', '')})" if ref
               else '260819_HumanMesh_wUpper_OMAKASE v4')
    n_hidden = sum(1 for r in meta if r.get('hidden'))
    ws['A3'] = (f'출처: Fusion 360 문서 {src_doc}, 알루미늄 바디 {len(meta)}개'
                + (f' (숨은 바디 {n_hidden}개 포함)' if n_hidden else '')
                + ' · 사진은 각 바디의 메시를 단독 렌더한 것')
    ws['A3'].font = Font(name=FONT, size=9, italic=True, color='7A828C')
    if measured:
        ws['A4'] = ('E열 기입값은 출력물 사진의 손글씨를 읽어 CAD 부품에 대조한 것 — '
                    '초록=확실 · 노랑=보통(A/B 짝 미확정) · 주황=불확실(실물 대조 필요) · 회색=알루미늄 가공품(CAD 질량 그대로). '
                    'F열이 46 %를 넘으면 빨강: PLA(1.24)/알루(2.70) 물리 상한 초과 = 오독 또는 오배정.')
        ws['A4'].font = Font(name=FONT, size=9, color='8A4B00')

    head = ['소속 링크', '사진', '파츠명', '질량 (알루 6061, g)',
            '측정 질량 (3D 프린트, g)', '측정/CAD', '비고']
    hr = 5
    for c, h in enumerate(head, start=1):
        cell = ws.cell(hr, c, h)
        cell.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[hr].height = 34

    widths = {'A': 22, 'B': 23, 'C': 30, 'D': 17, 'E': 19, 'F': 11, 'G': 58}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    n_img = 0
    for i, r in enumerate(meta):
        row = hr + 1 + i
        ws.row_dimensions[row].height = ROW_H
        ws.cell(row, 1, r['link']).alignment = Alignment(horizontal='left',
                                                         vertical='center', wrap_text=True)
        name = r['body'] if r['body'] == r['occ'].split(':')[0] else \
            f"{r['occ'].split(':')[0]} / {r['body']}"
        ws.cell(row, 3, name).alignment = Alignment(horizontal='left', vertical='center',
                                                    wrap_text=True)
        d = ws.cell(row, 4, round(r['mass_g'], 2))
        d.number_format = '#,##0.00'
        d.alignment = Alignment(horizontal='right', vertical='center')
        e = ws.cell(row, 5)
        e.fill = INPUT_FILL
        e.number_format = '#,##0.00'
        e.alignment = Alignment(horizontal='right', vertical='center')
        f = ws.cell(row, 6, f'=IF(E{row}="","",E{row}/D{row})')
        f.number_format = '0.0%'
        f.alignment = Alignment(horizontal='right', vertical='center')
        note = note_for(r)
        al = alu.get(r['path'])
        me = measured.get(r['path'])
        if al:
            # machined in aluminium after all: the CAD mass IS the mass, ratio 1.0 by
            # definition, and the PLA ceiling does not apply
            e.value = round(r['mass_g'], 2)
            e.fill = ALU_FILL
            note = '알루미늄(출력 아님) · ' + al['note']
        elif me:
            if me.get('g') is not None:
                e.value = me['g']
                e.fill = CONF_FILL.get(me['conf'], INPUT_FILL)
                if me['g'] / r['mass_g'] > PLA_CEILING:
                    f.font = Font(name=FONT, size=10, bold=True, color='C00000')
            bits = [f"사진{me['photo']} '{me['read']}'",
                    CONF_KO.get(me['conf'], '미기입')]
            if me.get('alt'):
                bits.append('대안: ' + me['alt'])
            if me.get('note'):
                bits.append(me['note'])
            note = ' · '.join(bits)
        g = ws.cell(row, 7, note)
        g.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        if g.value:
            g.fill = NOTE_FILL
        for c in range(1, 8):
            cell = ws.cell(row, c)
            cell.border = BORDER
            if not cell.font.bold and not (c == 6 and cell.font.color
                                           and cell.font.color.rgb == '00C00000'):
                cell.font = Font(name=FONT, size=10)
        img = r.get('img')
        if img and os.path.exists(img):
            pic = XLImage(img)
            pic.width, pic.height = IMG_W, IMG_H
            pic.anchor = f'B{row}'
            ws.add_image(pic)
            n_img += 1
        else:
            ws.cell(row, 2, '(사진 없음)').alignment = Alignment(horizontal='center',
                                                              vertical='center')

    last = hr + len(meta)
    tr = last + 1
    ws.cell(tr, 3, '합계').font = Font(name=FONT, size=10, bold=True)
    for col in ('D', 'E'):
        c = ws[f'{col}{tr}']
        c.value = f'=SUM({col}{hr + 1}:{col}{last})'
        c.number_format = '#,##0.00'
        c.font = Font(name=FONT, size=10, bold=True)
        c.border = BORDER
    ws.freeze_panes = f'A{hr + 1}'
    ws.auto_filter.ref = f'A{hr}:G{last}'

    # ---- per-link roll-up, fills in as the measurements arrive ----
    s2 = wb.create_sheet('링크별 집계')
    s2['A1'] = '링크별 합계 — E열에 측정값을 넣는 대로 자동으로 채워집니다'
    s2['A1'].font = Font(name=FONT, size=12, bold=True)
    h2 = ['소속 링크', '부품 수', 'CAD 알루 질량 (g)', '측정 질량 합 (g)', '측정 완료 수']
    for c, h in enumerate(h2, start=1):
        cell = s2.cell(3, c, h)
        cell.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    links = sorted({r['link'] for r in meta})
    rng = f"'알루미늄 부품'!$A${hr + 1}:$A${last}"
    for i, lk in enumerate(links):
        row = 4 + i
        s2.cell(row, 1, lk)
        s2.cell(row, 2, f'=COUNTIF({rng},A{row})')
        s2.cell(row, 3, f"=SUMIF({rng},A{row},'알루미늄 부품'!$D${hr + 1}:$D${last})")
        s2.cell(row, 4, f"=SUMIF({rng},A{row},'알루미늄 부품'!$E${hr + 1}:$E${last})")
        s2.cell(row, 5, f"=COUNTIFS({rng},A{row},'알루미늄 부품'!$E${hr + 1}:$E${last},\"<>\")")
        for c in range(1, 6):
            cell = s2.cell(row, c)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            if c >= 3:
                cell.number_format = '#,##0.0'
    tr2 = 4 + len(links)
    s2.cell(tr2, 1, '합계').font = Font(name=FONT, size=10, bold=True)
    for c in range(2, 6):
        cell = s2.cell(tr2, c, f'={get_column_letter(c)}4:{get_column_letter(c)}{tr2 - 1}')
        cell.value = f'=SUM({get_column_letter(c)}4:{get_column_letter(c)}{tr2 - 1})'
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.number_format = '#,##0.0' if c >= 3 else '0'
        cell.border = BORDER
    for col, w in {'A': 24, 'B': 10, 'C': 20, 'D': 20, 'E': 14}.items():
        s2.column_dimensions[col].width = w

    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f'{len(meta)} parts, {n_img} pictures embedded -> {out}')


if __name__ == '__main__':
    main()
