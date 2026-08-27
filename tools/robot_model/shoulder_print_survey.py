"""Weighing sheet for the printed parts of Torso2ShoulderP and Shoulder-Pitch2Roll.

Same shape as the earlier aluminium survey (docs/aluminium_parts_3dprint_masses.xlsx): one row
per part, its own picture, the mass CAD would give it at aluminium 6061 density as the
reference, and a yellow column to write the scale reading into. The ratio column then shows
how much lighter the printed part really is.

Why the aluminium reference at all, when these parts are printed: 6061 is the density the
parts were originally drawn in, so it is the common yardstick every earlier survey used. The
PLA column is what the model currently assumes, so the three columns together show CAD intent,
model assumption, and reality side by side.

Pictures need one mesh per body out of Fusion. If the render directory is missing the sheet is
still written, with the picture column empty and a note - the numbers are what the weighing
needs, and the pictures can be added by re-running once the connector is back.

  shoulder_print_survey.py [out.xlsx]        (mjlab .venv python)
"""
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = '/home/syaro/MikuchanRemote/Human-Pygmalion'
BODIES = '/home/syaro/pyg_fea/fusion/bodies_v4_printed.json'
IMG_DIR = '/home/syaro/pyg_fea/fusion/shoulder_parts/img'
OUT = sys.argv[1] if len(sys.argv) > 1 else f'{REPO}/docs/shoulder_parts_3dprint_masses.xlsx'

SCOPES = ('Torso2ShoulderP', 'Shoulder-Pitch2Roll')
RHO_AL = 2.70      # g/cm3, aluminium 6061 - the density the parts were drawn in
RHO_PLA = 0.887    # g/cm3, the printed density the model currently assumes
IMG_W, IMG_H, ROW_H = 156, 117, 92

HDR = PatternFill('solid', fgColor='DDE3EA')
ENTRY = PatternFill('solid', fgColor='FFF6CC')       # the only cells a human fills in
WARN = PatternFill('solid', fgColor='F8D7C0')
THIN = Side(style='thin', color='B0B7C0')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def collect():
    B = json.load(open(BODIES, encoding='utf-8'))
    rows = []
    for path, d in B.items():
        if not any(s in path for s in SCOPES):
            continue
        if 'Screw' in path or 'Washer' in path:          # fasteners are bought, not printed
            continue
        name = path.split('::')[-1]
        comp = next(s for s in SCOPES if s in path)
        vol = d['v']
        rows.append({
            'comp': comp, 'name': name, 'path': path,
            'vol': vol, 'cad_g': d['m'] * 1000.0,
            'rho': (d['m'] * 1000.0 / vol) if vol else 0.0,
            'al_g': vol * RHO_AL, 'pla_g': vol * RHO_PLA,
            'mat': d.get('mat', '?'),
            'bbox': d.get('bb'),
        })
    rows.sort(key=lambda r: (r['comp'], -r['vol']))
    return rows


def main():
    rows = collect()
    wb = Workbook()
    ws = wb.active
    ws.title = '어깨 출력물'

    ws['A1'] = 'Pygmalion 어깨 출력물 — 실측 질량 조사'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = ('노란 칸(G열)에만 입력하세요. 저울로 잰 값을 그램 숫자로 — 예: 84.2')
    ws['A3'] = (f'대상: {" · ".join(SCOPES)} 안의 출력 부품 (나사·와셔 제외). '
                f'알루미늄 기준 = 부피 x {RHO_AL} g/cm3 (6061), PLA 기준 = 부피 x {RHO_PLA} g/cm3')
    ws['A4'] = ('출처: Fusion 360 문서 260819_HumanMesh_wUpper_URDFexport_v22 에서 뜬 부피. '
                '부피는 재질과 무관하므로 기준 질량은 모두 이 부피에서 계산된 값입니다.')
    for r in (2, 3, 4):
        ws[f'A{r}'].font = Font(size=9, color='666666')

    head = ['컴포넌트', '사진', '파츠명', '부피 (cm³)',
            f'알루 6061 기준 (g)', f'PLA {RHO_PLA} 기준 (g)',
            '측정 질량 (g)', '측정/알루', '측정/PLA', 'CAD 현재값 (g)', '비고']
    for c, h in enumerate(head, 1):
        cell = ws.cell(6, c, h)
        cell.font = Font(bold=True)
        cell.fill = HDR
        cell.border = BOX
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    have_img = os.path.isdir(IMG_DIR)
    r = 7
    for row in rows:
        ws.cell(r, 1, row['comp']).border = BOX
        ws.cell(r, 2, '').border = BOX
        ws.cell(r, 3, row['name']).border = BOX
        ws.cell(r, 4, round(row['vol'], 3)).border = BOX
        ws.cell(r, 5, round(row['al_g'], 2)).border = BOX
        ws.cell(r, 6, round(row['pla_g'], 2)).border = BOX
        e = ws.cell(r, 7, None)
        e.fill = ENTRY
        e.border = BOX
        ws.cell(r, 8, f'=IF(G{r}="","",G{r}/E{r})').border = BOX
        ws.cell(r, 9, f'=IF(G{r}="","",G{r}/F{r})').border = BOX
        ws.cell(r, 10, round(row['cad_g'], 2)).border = BOX
        # A part still carrying the CAD default density is worth flagging on the sheet itself:
        # it means the model is currently using a mass that nobody chose.
        note = ''
        if row['rho'] > 3.0:
            note = f'⚠ CAD가 아직 금속 밀도({row["rho"]:.2f} g/cm³) — 출력물이면 모델 질량이 과대'
            ws.cell(r, 10).fill = WARN
        ws.cell(r, 11, note).border = BOX
        img = f'{IMG_DIR}/{row["name"]}.png'
        if have_img and os.path.exists(img):
            pic = XLImage(img)
            pic.width, pic.height = IMG_W, IMG_H
            ws.add_image(pic, f'B{r}')
            ws.row_dimensions[r].height = ROW_H
        r += 1

    n = r - 1
    ws.cell(r + 1, 3, '합계').font = Font(bold=True)
    for col, letter in ((4, 'D'), (5, 'E'), (6, 'F'), (7, 'G'), (10, 'J')):
        c = ws.cell(r + 1, col, f'=SUM({letter}7:{letter}{n})')
        c.font = Font(bold=True)
        c.border = BOX

    for col, w in (('A', 21), ('B', 23), ('C', 30), ('D', 12), ('E', 16),
                   ('F', 15), ('G', 15), ('H', 11), ('I', 11), ('J', 15), ('K', 46)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A7'

    if not have_img:
        ws.cell(r + 3, 1, '※ 사진 열은 비어 있습니다 — Fusion 커넥터의 스크립트 실행이 '
                          '중단되어 부품별 메쉬를 받아오지 못했습니다. 복구 후 이 스크립트를 '
                          '다시 실행하면 숫자는 그대로 두고 사진만 채워집니다.').font = \
            Font(size=9, color='B04A2A')

    wb.save(OUT)
    print(f'wrote {OUT}  ({n - 6} parts)')
    print(f"{'component':22s} {'part':30s} {'vol':>8s} {'alu g':>9s} {'PLA g':>9s} {'CAD g':>9s}")
    for row in rows:
        flag = '  <- CAD still metal' if row['rho'] > 3.0 else ''
        print(f"{row['comp']:22s} {row['name'][:30]:30s} {row['vol']:8.2f} "
              f"{row['al_g']:9.2f} {row['pla_g']:9.2f} {row['cad_g']:9.2f}{flag}")
    tot_al = sum(x['al_g'] for x in rows)
    tot_pla = sum(x['pla_g'] for x in rows)
    tot_cad = sum(x['cad_g'] for x in rows)
    tot_vol = sum(x['vol'] for x in rows)
    print(f"{'':22s} {'TOTAL':30s} {tot_vol:8.2f} {tot_al:9.2f} {tot_pla:9.2f} {tot_cad:9.2f}")
    print(f'CAD 현재값이 PLA 기준보다 {tot_cad - tot_pla:+.1f} g 무겁습니다.')


if __name__ == '__main__':
    main()
